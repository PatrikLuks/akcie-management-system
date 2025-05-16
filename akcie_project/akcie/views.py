import csv
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse, FileResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg
from django.db.models.functions import TruncMonth
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from django.contrib.auth.decorators import login_required, user_passes_test
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from io import BytesIO
from django.urls import reverse
from django.core.mail import EmailMessage
from django.utils.timezone import now
from django.utils import timezone
from datetime import datetime, timedelta
import os
import zipfile
import requests
import yfinance as yf
from .models import Akcie, Transakce, Dividenda, Aktivita, CustomUser, AuditLog, Klient, Portfolio
from .forms import AkcieForm, TransakceForm, DividendaForm, CustomUserForm, KlientForm, PortfolioForm
from .utils_pdf import add_luxury_branding
from functools import wraps
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
import pandas as pd

API_URL = 'https://example.com/api/akcie'  # Nahraďte skutečnou URL API
TREND_API_URL = 'https://api.example.com/trends'  # Skutečná URL API

# --- Role-based access utility dekorátory ---
def is_admin(user):
    return user.groups.filter(name='Admin').exists() or user.is_superuser

def is_poradce(user):
    return user.groups.filter(name='Poradce').exists()

def is_klient(user):
    return user.groups.filter(name='Klient').exists()

def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not is_admin(request.user):
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('Přístup pouze pro administrátory.')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def poradce_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not (is_admin(request.user) or is_poradce(request.user)):
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('Přístup pouze pro poradce.')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def klient_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not (is_admin(request.user) or is_poradce(request.user) or is_klient(request.user)):
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('Přístup pouze pro klienty.')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def fetch_akcie_data():
    """
    Získá data o akciích z externího API.
    """
    response = requests.get(API_URL)
    if response.status_code == 200:
        return response.json()
    else:
        return []

def fetch_hot_investments():
    """
    Získá 10 nejlepších investic podle trendů pomocí knihovny yfinance.
    """
    try:
        # Example: Fetch data for popular tickers
        tickers = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'META', 'NVDA', 'BRK-B', 'JNJ', 'V']
        investments = []
        for idx, ticker in enumerate(tickers):
            stock = yf.Ticker(ticker)
            info = stock.info
            investments.append({
                'id': idx + 1,  # Simulated ID for demonstration purposes
                'nazev': info.get('shortName', ticker),
                'cena_za_ks': info.get('regularMarketPrice', 'N/A'),
                'hodnota': info.get('marketCap', 'N/A')
            })
        return investments[:10]  # Vrátí prvních 10 investic
    except Exception as e:
        print(f"Chyba při získávání dat z yfinance: {e}")
        return []

def log_aktivita(akce, uzivatel=None):
    """
    Loguje uživatelské aktivity do databáze.
    :param akce: Popis akce, která byla provedena.
    :param uzivatel: Uživatelské jméno (volitelné).
    """
    Aktivita.objects.create(akce=akce, uzivatel=uzivatel)

# Přidání pomocné funkce pro filtrování akcií

def filter_akcie(query):
    """Filtruje akcie podle zadaného dotazu."""
    return Akcie.objects.filter(
        Q(nazev__icontains=query) |
        Q(pocet_ks__icontains=query) |
        Q(cena_za_kus__icontains=query)
    )

# Pomocná funkce pro převod měny na CZK

def convert_to_czk(amount, from_currency):
    if from_currency == 'CZK':
        return float(amount)
    try:
        r = requests.get('https://www.cnb.cz/en/financial-markets/foreign-exchange-market/exchange-rate-fixing/daily.txt')
        if r.status_code == 200:
            lines = r.text.split('\n')
            for line in lines:
                # Opravdu přesné porovnání měny (např. 'USD|1|...')
                if line.startswith(f'{from_currency}|'):
                    parts = line.split('|')
                    kurz = float(parts[4].replace(',', '.'))
                    print(f"[DEBUG] CNB {from_currency}/CZK kurz použit: {kurz}")
                    return float(amount) * kurz
        print(f"[WARNING] Kurz pro {from_currency}/CZK není dostupný, použit fallback 23.0")
        return float(amount) * 23.0
    except Exception as e:
        print(f"Chyba při převodu měny: {e}")
        return float(amount) * 23.0

def convert_to_czk_api(request):
    try:
        cena = float(request.GET.get('cena', 0))
        mena = request.GET.get('mena', 'CZK')
        from .views import convert_to_czk
        czk = round(convert_to_czk(cena, mena), 2)
        return JsonResponse({'czk': czk})
    except Exception as e:
        return JsonResponse({'czk': None, 'error': str(e)})

@login_required
@poradce_required
def klient_list(request):
    klienti = Klient.objects.all()
    return render(request, 'akcie/klient_list.html', {'klienti': klienti})

@login_required
@admin_required
def klient_list_admin(request):
    poradci = CustomUser.objects.filter(groups__name='Poradce')
    poradce_id = request.GET.get('poradce')
    klienti = Klient.objects.all()
    if poradce_id:
        klienti = klienti.filter(poradce_id=poradce_id)
    return render(request, 'akcie/klient_list_admin.html', {'klienti': klienti, 'poradci': poradci, 'poradce_id': poradce_id})

@login_required
@poradce_required
def klient_create(request):
    if request.method == 'POST':
        form = KlientForm(request.POST)
        if form.is_valid():
            klient = form.save(commit=False)
            klient.save()
            return redirect('klient_list')
    else:
        form = KlientForm()
    return render(request, 'akcie/klient_form.html', {'form': form})

@login_required
@poradce_required
def klient_update(request, pk):
    klient = get_object_or_404(Klient, pk=pk)
    if request.method == 'POST':
        form = KlientForm(request.POST, instance=klient)
        if form.is_valid():
            form.save()
            return redirect('klient_list')
    else:
        form = KlientForm(instance=klient)
    return render(request, 'akcie/klient_form.html', {'form': form})

@login_required
@admin_required
def klient_delete(request, pk):
    klient = get_object_or_404(Klient, pk=pk)
    if request.method == 'POST':
        klient.delete()
        return redirect('klient_list')
    return render(request, 'akcie/klient_confirm_delete.html', {'klient': klient})

@login_required
@poradce_required
def klient_report_pdf(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id, poradce=request.user)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_{klient.jmeno}_{klient.prijmeni}.pdf"'
    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 750, f"Report klienta: {klient.jmeno} {klient.prijmeni}")
    p.setFont("Helvetica", 12)
    p.drawString(50, 730, f"Email: {klient.email}")
    p.drawString(50, 715, f"Telefon: {klient.telefon}")
    p.drawString(50, 700, f"Adresa: {klient.adresa}")
    y = 670
    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "Portfolia:")
    y -= 20
    p.setFont("Helvetica", 12)
    for portfolio in klient.portfolia.all():
        p.drawString(60, y, f"- {portfolio.nazev}: {portfolio.popis}")
        y -= 15
        if y < 100:
            p.showPage()
            y = 750
    p.showPage()
    p.save()
    return response

@login_required
@poradce_required
def klient_report_send_email(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id, poradce=request.user)
    # Vygeneruj PDF do paměti
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 750, f"Report klienta: {klient.jmeno} {klient.prijmeni}")
    p.setFont("Helvetica", 12)
    p.drawString(50, 730, f"Email: {klient.email}")
    p.drawString(50, 715, f"Telefon: {klient.telefon}")
    p.drawString(50, 700, f"Adresa: {klient.adresa}")
    y = 670
    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "Portfolia:")
    y -= 20
    p.setFont("Helvetica", 12)
    for portfolio in klient.portfolia.all():
        p.drawString(60, y, f"- {portfolio.nazev}: {portfolio.popis}")
        y -= 15
        if y < 100:
            p.showPage()
            y = 750
    p.showPage()
    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    # Odeslání e-mailem
    subject = f"Report klienta {klient.jmeno} {klient.prijmeni}"
    body = f"Dobrý den,\nv příloze naleznete report klienta {klient.jmeno} {klient.prijmeni}."
    email = EmailMessage(subject, body, to=[klient.email])
    email.attach(f"report_{klient.jmeno}_{klient.prijmeni}.pdf", pdf, 'application/pdf')
    email.send()
    messages.success(request, f'Report byl odeslán na e-mail: {klient.email}')
    return redirect('poradce_dashboard')

@login_required
@poradce_required
def portfolio_list(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id)
    portfolia = klient.portfolia.all()
    return render(request, 'akcie/portfolio_list.html', {'klient': klient, 'portfolia': portfolia})

@login_required
@poradce_required
def portfolio_create(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id)
    if request.method == 'POST':
        form = PortfolioForm(request.POST)
        if form.is_valid():
            portfolio = form.save(commit=False)
            portfolio.klient = klient
            portfolio.save()
            return redirect('portfolio_list', klient_id=klient.id)
    else:
        form = PortfolioForm()
    return render(request, 'akcie/portfolio_form.html', {'form': form, 'klient': klient})

@login_required
@poradce_required
def portfolio_update(request, klient_id, pk):
    klient = get_object_or_404(Klient, pk=klient_id)
    portfolio = get_object_or_404(Portfolio, pk=pk, klient=klient)
    if request.method == 'POST':
        form = PortfolioForm(request.POST, instance=portfolio)
        if form.is_valid():
            form.save()
            return redirect('portfolio_list', klient_id=klient.id)
    else:
        form = PortfolioForm(instance=portfolio)
    return render(request, 'akcie/portfolio_form.html', {'form': form, 'klient': klient})

@login_required
@admin_required
def portfolio_delete(request, klient_id, pk):
    klient = get_object_or_404(Klient, pk=klient_id)
    portfolio = get_object_or_404(Portfolio, pk=pk, klient=klient)
    if request.method == 'POST':
        portfolio.delete()
        return redirect('portfolio_list', klient_id=klient.id)
    return render(request, 'akcie/portfolio_confirm_delete.html', {'portfolio': portfolio, 'klient': klient})

@login_required
@poradce_required
def export_klienti_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="klienti.csv"'
    writer = csv.writer(response)
    writer.writerow(['Jméno', 'Příjmení', 'Email', 'Telefon', 'Adresa', 'Poznámka'])
    for klient in Klient.objects.all():
        writer.writerow([
            klient.jmeno,
            klient.prijmeni,
            klient.email,
            klient.telefon,
            klient.adresa,
            klient.poznamka
        ])
    writer.writerow([])
    writer.writerow(["--- Luxusní export pro Finanční Poradce Premium ---"])
    log_aktivita("Export klientů do CSV", request.user.username)
    return response

@login_required
@poradce_required
def export_klienti_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="klienti.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Klienti"
    headers = ['Jméno', 'Příjmení', 'Email', 'Telefon', 'Adresa', 'Poznámka']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="007bff")
    for klient in Klient.objects.all():
        ws.append([
            klient.jmeno,
            klient.prijmeni,
            klient.email,
            klient.telefon,
            klient.adresa,
            klient.poznamka
        ])
    ws.append([])
    ws.append(["--- Luxusní export pro Finanční Poradce Premium ---"])
    log_aktivita("Export klientů do Excelu", request.user.username)
    wb.save(response)
    return response

@login_required
@poradce_required
def export_klienti_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="klienti_report.pdf"'
    p = canvas.Canvas(response, pagesize=letter)
    from .utils_pdf import add_luxury_branding
    add_luxury_branding(p)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, "Report klientů")
    p.setFont("Helvetica", 12)
    y = 720
    for klient in Klient.objects.all():
        p.drawString(100, y, f"{klient.jmeno} {klient.prijmeni} | {klient.email} | {klient.telefon}")
        y -= 20
        if y < 50:
            p.showPage()
            add_luxury_branding(p)
            p.setFont("Helvetica", 12)
            y = 750
    p.showPage()
    p.save()
    log_aktivita("Export klientů do PDF", request.user.username)
    return response

@login_required
@poradce_required
def export_portfolia_csv(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="portfolia_{klient.jmeno}_{klient.prijmeni}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Název', 'Popis', 'Datum vytvoření'])
    for portfolio in klient.portfolia.all():
        writer.writerow([
            portfolio.nazev,
            portfolio.popis,
            portfolio.datum_vytvoreni.strftime('%d.%m.%Y %H:%M')
        ])
    writer.writerow([])
    writer.writerow(["--- Luxusní export pro Finanční Poradce Premium ---"])
    log_aktivita(f"Export portfolií klienta {klient}", request.user.username)
    return response

@login_required
@poradce_required
def export_portfolia_excel(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="portfolia_{klient.jmeno}_{klient.prijmeni}.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolia"
    headers = ['Název', 'Popis', 'Datum vytvoření']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="007bff")
    for portfolio in klient.portfolia.all():
        ws.append([
            portfolio.nazev,
            portfolio.popis,
            portfolio.datum_vytvoreni.strftime('%d.%m.%Y %H:%M')
        ])
    ws.append([])
    ws.append(["--- Luxusní export pro Finanční Poradce Premium ---"])
    log_aktivita(f"Export portfolií klienta {klient}", request.user.username)
    wb.save(response)
    return response

@login_required
@poradce_required
def export_portfolia_pdf(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="portfolia_{klient.jmeno}_{klient.prijmeni}.pdf"'
    p = canvas.Canvas(response, pagesize=letter)
    from .utils_pdf import add_luxury_branding
    add_luxury_branding(p)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, f"Report portfolií klienta {klient.jmeno} {klient.prijmeni}")
    p.setFont("Helvetica", 12)
    y = 720
    for portfolio in klient.portfolia.all():
        p.drawString(100, y, f"{portfolio.nazev} | {portfolio.popis} | {portfolio.datum_vytvoreni.strftime('%d.%m.%Y %H:%M')}")
        y -= 20
        if y < 50:
            p.showPage()
            add_luxury_branding(p)
            p.setFont("Helvetica", 12)
            y = 750
    p.showPage()
    p.save()
    log_aktivita(f"Export portfolií klienta {klient}", request.user.username)
    return response

@login_required
def klient_report_pdf(request, klient_id):
    klient = get_object_or_404(Klient, pk=klient_id, poradce=request.user)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_{klient.jmeno}_{klient.prijmeni}.pdf"'
    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 750, f"Report klienta: {klient.jmeno} {klient.prijmeni}")
    p.setFont("Helvetica", 12)
    p.drawString(50, 730, f"Email: {klient.email}")
    p.drawString(50, 715, f"Telefon: {klient.telefon}")
    p.drawString(50, 700, f"Adresa: {klient.adresa}")
    y = 670
    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y, "Portfolia:")
    y -= 20
    p.setFont("Helvetica", 12)
    for portfolio in klient.portfolia.all():
        p.drawString(60, y, f"- {portfolio.nazev}: {portfolio.popis}")
        y -= 15
        if y < 100:
            p.showPage()
            y = 750
    p.showPage()
    p.save()
    return response

@login_required
def index(request):
    """
    Úvodní stránka zobrazující akcie vybrané uživatelem a grafy pro finanční poradce.
    """
    user_stocks = Akcie.objects.filter(user=request.user)

    # Oprava: vždy získat aktuální cenu a měnu z yfinance pro každou akcii
    user_stocks_czk = []
    for akcie in user_stocks:
        ticker = getattr(akcie, 'ticker', None) or akcie.nazev
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            current_price = float(info.get('regularMarketPrice', akcie.cena_za_kus))
            currency = info.get('currency', 'CZK')
        except Exception:
            current_price = akcie.cena_za_kus
            currency = 'CZK'
        current_price_czk = convert_to_czk(current_price, currency)
        hodnota_czk = akcie.pocet_ks * current_price_czk
        nakup_czk = float(akcie.nakup)  # OPRAVA: vždy převést na float
        zisk_ztrata_czk = hodnota_czk - nakup_czk
        user_stocks_czk.append({
            'id': akcie.id,
            'nazev': akcie.nazev,
            'pocet_ks': akcie.pocet_ks,
            'cena_za_kus': round(current_price_czk, 2),
            'hodnota': round(hodnota_czk, 2),
            'zisk_ztrata': round(zisk_ztrata_czk, 2),
        })

    # --- NOVÁ LOGIKA: Vývoj hodnoty portfolia v čase podle transakcí a historických cen ---
    from collections import defaultdict
    import yfinance as yf
    from datetime import date, timedelta
    import calendar

    # 1. Zjisti všechny tickery v portfoliu uživatele
    tickers = set()
    for akcie in user_stocks:
        ticker = getattr(akcie, 'ticker', None)
        if ticker:
            tickers.add(ticker)
        else:
            tickers.add(akcie.nazev)  # fallback, pokud není ticker

    # 2. Zjisti všechny měsíce, kdy uživatel držel nějakou akcii
    transakce = Transakce.objects.filter(akcie__user=request.user).order_by('datum')
    if not transakce.exists():
        history_labels = []
        history_values = []
    else:
        first = transakce.first().datum.replace(day=1)
        last = date.today().replace(day=1)
        months = []
        d = first
        while d <= last:
            months.append(d)
            if d.month == 12:
                d = d.replace(year=d.year+1, month=1)
            else:
                d = d.replace(month=d.month+1)
        history_labels = [d.strftime('%Y-%m') for d in months]
        history_values = []
        for d in months:
            total = 0
            for akcie in user_stocks:
                # Kolik kusů akcie uživatel v daném měsíci držel?
                last_day = d.replace(day=calendar.monthrange(d.year, d.month)[1])
                trans = Transakce.objects.filter(akcie=akcie, datum__lte=last_day)
                ks = 0
                for t in trans:
                    if t.typ == 'nákup':
                        ks += t.mnozstvi
                    elif t.typ == 'prodej':
                        ks -= t.mnozstvi
                if ks <= 0:
                    continue
                # Získej historickou cenu akcie k poslednímu dni měsíce
                ticker = getattr(akcie, 'ticker', None) or akcie.nazev
                try:
                    stock = yf.Ticker(ticker)
                    hist = stock.history(start=last_day.strftime('%Y-%m-%d'), end=(last_day+timedelta(days=1)).strftime('%Y-%m-%d'))
                    if not hist.empty:
                        cena = float(hist.iloc[0]['Close'])
                    else:
                        cena = float(stock.info.get('regularMarketPrice', 0))
                except Exception:
                    cena = 0
                total += ks * cena
            history_values.append(round(total, 2))

    # --- Ostatní grafy beze změny ---
    dividend_history = (
        Dividenda.objects.filter(akcie__user=request.user)
        .values('datum')
        .order_by('datum')
        .annotate(total_dividend=Sum('castka'))
    )
    dividend_labels = [str(item['datum']) for item in dividend_history]
    dividend_values = [float(item['total_dividend']) for item in dividend_history]

    stock_distribution = (
        Akcie.objects.filter(user=request.user)
        .values('nazev')
        .annotate(total_value=Sum('hodnota'))
    )
    dist_labels = [item['nazev'] for item in stock_distribution]
    dist_values = [float(item['total_value']) for item in stock_distribution]

    context = {
        'user_stocks': user_stocks_czk,
        'history_labels': history_labels if history_labels is not None else [],
        'history_values': history_values if history_values is not None else [],
        'dividend_labels': dividend_labels if dividend_labels is not None else [],
        'dividend_values': dividend_values if dividend_values is not None else [],
        'dist_labels': dist_labels if dist_labels is not None else [],
        'dist_values': dist_values if dist_values is not None else [],
    }
    return render(request, 'akcie/index.html', context)

@login_required
def akcie_list(request):
    import yfinance as yf
    from decimal import Decimal
    query = request.GET.get('q')
    akcie_qs = filter_akcie(query) if query else Akcie.objects.all()
    akcie_list = []
    for akcie in akcie_qs:
        ticker = getattr(akcie, 'ticker', None) or akcie.nazev
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            current_price = float(info.get('regularMarketPrice', akcie.cena_za_kus))
            currency = info.get('currency', akcie.mena or 'CZK')
        except Exception as e:
            print(f"[DEBUG] yfinance error for {ticker}: {e}")
            current_price = float(akcie.cena_za_kus)
            currency = akcie.mena or 'CZK'
        try:
            current_price_czk = float(convert_to_czk(current_price, currency))
        except Exception as e:
            print(f"[DEBUG] convert_to_czk error for {ticker}: {e}")
            current_price_czk = current_price
        hodnota_czk = float(akcie.pocet_ks) * current_price_czk
        if akcie.mena == 'CZK' or not akcie.mena:
            nakup_czk = float(akcie.nakup)
        else:
            nakup_czk = float(akcie.nakup)
        zisk_ztrata_czk = hodnota_czk - nakup_czk
        if ticker and ticker.lower() == 'aapl':
            print(f"[DEBUG] AAPL: cena={current_price} {currency}, cena_czk={current_price_czk}, hodnota_czk={hodnota_czk}, nakup_czk={nakup_czk}, zisk_ztrata_czk={zisk_ztrata_czk}")
        akcie_list.append({
            'id': akcie.id,
            'nazev': akcie.nazev,
            'ticker': akcie.ticker,
            'mena': currency,
            'pocet_ks': akcie.pocet_ks,
            'datum': akcie.datum,
            'hodnota': round(hodnota_czk, 2),
            'zisk_ztrata': round(zisk_ztrata_czk, 2),
        })
    return render(request, 'akcie/akcie_list.html', {'akcie': akcie_list})

# Refaktorování akcie_detail

def get_akcie_detail_context(akcie):
    """Vrací kontext pro detail akcie."""
    transakce = Transakce.objects.filter(akcie=akcie)
    dividendy = Dividenda.objects.filter(akcie=akcie)
    total_dividendy = dividendy.aggregate(Sum('castka'))['castka__sum'] or 0
    return {
        'akcie': akcie,
        'transakce': transakce,
        'dividendy': dividendy,
        'total_dividendy': total_dividendy,
    }

def akcie_detail(request, pk):
    """Zobrazuje detail akcie."""
    akcie = get_object_or_404(Akcie, pk=pk)
    context = get_akcie_detail_context(akcie)
    return render(request, 'akcie/akcie_detail.html', context)

@login_required
@poradce_required
def akcie_create(request):
    if request.method == 'POST':
        ticker = request.POST.get('ticker')
        nazev = request.POST.get('nazev')
        datum = request.POST.get('datum')
        cas = request.POST.get('cas') or '00:00'
        pocet_ks = int(request.POST.get('pocet_ks'))
        cena_za_kus = 0
        current_price = 0
        currency = 'CZK'
        try:
            stock = yf.Ticker(ticker)
            datum_dt = datetime.strptime(datum, "%Y-%m-%d")
            end_dt = datum_dt + timedelta(days=1)
            hist = stock.history(start=datum_dt.strftime("%Y-%m-%d"), end=end_dt.strftime("%Y-%m-%d"))
            info = stock.info
            currency = info.get('currency', 'CZK')
            if not hist.empty:
                cena_za_kus = float(hist.iloc[0]['Close'])
            else:
                for i in range(1, 7):
                    prev_dt = datum_dt - timedelta(days=i)
                    prev_end = prev_dt + timedelta(days=1)
                    prev_hist = stock.history(start=prev_dt.strftime("%Y-%m-%d"), end=prev_end.strftime("%Y-%m-%d"))
                    if not prev_hist.empty:
                        cena_za_kus = float(prev_hist.iloc[0]['Close'])
                        break
                else:
                    cena_za_kus = float(info.get('regularMarketPrice', 0))
            current_price = float(info.get('regularMarketPrice', cena_za_kus))
        except Exception as e:
            print(f"[ERROR] yfinance fetch: {e}")
            cena_za_kus = 0
            current_price = 0
            currency = 'CZK'
        cena_za_kus_czk = convert_to_czk(cena_za_kus, currency)
        current_price_czk = convert_to_czk(current_price, currency)
        hodnota = pocet_ks * current_price_czk
        nakup = pocet_ks * cena_za_kus_czk
        zisk_ztrata = hodnota - nakup
        dividenda = hodnota * 0.05
        akcie = Akcie.objects.create(
            user=request.user,
            nazev=nazev,
            datum=datum,
            cas=cas,
            pocet_ks=pocet_ks,
            cena_za_kus=cena_za_kus_czk,
            hodnota=hodnota,
            nakup=nakup,
            zisk_ztrata=zisk_ztrata,
            dividenda=dividenda,
            ticker=ticker,
            mena=currency
        )
        return redirect('akcie_list')
    else:
        form = AkcieForm()
    return render(request, 'akcie/akcie_form.html', {'form': form})

@login_required
@poradce_required
def akcie_update(request, pk):
    akcie = get_object_or_404(Akcie, pk=pk)
    if request.method == 'POST':
        form = AkcieForm(request.POST, instance=akcie)
        if form.is_valid():
            form.save()
            return redirect('akcie_list')
    else:
        form = AkcieForm(instance=akcie)
    return render(request, 'akcie/akcie_form.html', {'form': form})

@login_required
@admin_required
def akcie_delete(request, pk):
    akcie = get_object_or_404(Akcie, pk=pk)
    if request.method == 'POST':
        akcie.delete()
        return redirect('akcie_list')
    return render(request, 'akcie/akcie_confirm_delete.html', {'akcie': akcie})

@login_required
def transakce_list(request):
    query = request.GET.get('q')
    if query:
        transakce = Transakce.objects.filter(
            Q(akcie__nazev__icontains=query) |
            Q(typ__icontains=query) |
            Q(mnozstvi__icontains=query) |
            Q(cena__icontains=query)
        )
    else:
        transakce = Transakce.objects.all()
    return render(request, 'akcie/transakce_list.html', {'transakce': transakce})

def transakce_detail(request, pk):
    transakce = Transakce.objects.get(pk=pk)
    return render(request, 'akcie/transakce_detail.html', {'transakce': transakce})

@login_required
@poradce_required
def transakce_create(request):
    """
    Vytvoření nové transakce s daty z API.
    """
    akcie_data = fetch_akcie_data()
    if not akcie_data:
        akcie_data = [{'nazev': 'Žádná data nejsou dostupná'}]  # Výchozí hodnota

    if request.method == 'POST':
        form = TransakceForm(request.POST)
        if form.is_valid():
            transakce = form.save(commit=False)
            selected_akcie = next((akcie for akcie in akcie_data if akcie['nazev'] == form.cleaned_data['akcie']), None)
            if selected_akcie:
                transakce.cena = selected_akcie.get('cena_za_ks', 0) * form.cleaned_data['mnozstvi']
                transakce.hodnota = selected_akcie.get('hodnota', 0)
                transakce.zisk_ztrata = selected_akcie.get('zisk', 0)
                transakce.dividenda = selected_akcie.get('dividenda', 0)
            transakce.save()
            return redirect('transakce_list')
    else:
        form = TransakceForm()
    return render(request, 'akcie/transakce_form.html', {'form': form, 'akcie_data': akcie_data})

@login_required
@poradce_required
def transakce_update(request, pk):
    transakce = get_object_or_404(Transakce, pk=pk)
    if request.method == 'POST':
        form = TransakceForm(request.POST, instance=transakce)
        if form.is_valid():
            form.save()
            return redirect('transakce_list')
    else:
        form = TransakceForm(instance=transakce)
    return render(request, 'akcie/transakce_form.html', {'form': form})

@login_required
@admin_required
def transakce_delete(request, pk):
    transakce = get_object_or_404(Transakce, pk=pk)
    if request.method == 'POST':
        transakce.delete()
        return redirect('transakce_list')
    return render(request, 'akcie/transakce_confirm_delete.html', {'transakce': transakce})

@login_required
def dividenda_list(request):
    query = request.GET.get('q')
    akcie_filter = request.GET.get('akcie')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    dividendy = Dividenda.objects.all()
    if query:
        dividendy = dividendy.filter(
            Q(akcie__nazev__icontains=query) |
            Q(castka__icontains=query)
        )
    if akcie_filter:
        dividendy = dividendy.filter(akcie__id=akcie_filter)
    if date_from:
        dividendy = dividendy.filter(datum__gte=date_from)
    if date_to:
        dividendy = dividendy.filter(datum__lte=date_to)

    # Souhrnné statistiky
    from django.db.models import Avg, Max, Min, Count, Sum
    total_dividendy = dividendy.aggregate(Sum('castka'))['castka__sum'] or 0
    avg_dividenda = dividendy.aggregate(Avg('castka'))['castka__avg'] or 0
    count_dividendy = dividendy.count()
    nejblizsi = dividendy.filter(datum__gte=now().date()).order_by('datum').first()
    nejblizsi_vyplata = nejblizsi.datum if nejblizsi else None
    nejblizsi_castka = nejblizsi.castka if nejblizsi else None

    # Vývoj dividend v čase (agregace podle měsíce)
    from django.db.models.functions import TruncMonth
    timeline = (
        dividendy
        .annotate(month=TruncMonth('datum'))
        .values('month')
        .order_by('month')
        .annotate(total=Sum('castka'))
    )
    timeline_labels = [str(item['month']) for item in timeline]
    timeline_values = [float(item['total']) for item in timeline]

    # Rozložení dividend podle akcií
    rozlozeni = (
        dividendy
        .values('akcie__nazev')
        .annotate(total=Sum('castka'))
        .order_by('-total')
    )
    rozlozeni_labels = [item['akcie__nazev'] for item in rozlozeni]
    rozlozeni_values = [float(item['total']) for item in rozlozeni]

    # Pro filtr akcií
    akcie_list = Akcie.objects.all()

    sum_dividendy = dividendy.aggregate(sum=Sum('castka'))['sum'] or 0
    context = {
        'dividendy': dividendy,
        'total_dividendy': total_dividendy,
        'avg_dividenda': avg_dividenda,
        'count_dividendy': count_dividendy,
        'nejblizsi_vyplata': nejblizsi_vyplata,
        'nejblizsi_castka': nejblizsi_castka,
        'timeline_labels': timeline_labels,
        'timeline_values': timeline_values,
        'rozlozeni_labels': rozlozeni_labels,
        'rozlozeni_values': rozlozeni_values,
        'today': now().date(),
        'request': request,
        'akcie_list': akcie_list,
        'date_from': date_from,
        'date_to': date_to,
        'akcie_filter': akcie_filter,
        'sum_dividendy': sum_dividendy,
    }
    return render(request, 'akcie/dividenda_list.html', context)

def dividenda_detail(request, pk):
    dividenda = Dividenda.objects.get(pk=pk)
    return render(request, 'akcie/dividenda_detail.html', {'dividenda': dividenda})

@login_required
@poradce_required
def dividenda_create(request):
    if request.method == 'POST':
        form = DividendaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dividenda_list')
    else:
        form = DividendaForm()
    return render(request, 'akcie/dividenda_form.html', {'form': form})

@login_required
@poradce_required
def dividenda_update(request, pk):
    dividenda = get_object_or_404(Dividenda, pk=pk)
    if request.method == 'POST':
        form = DividendaForm(request.POST, instance=dividenda)
        if form.is_valid():
            form.save()
            return redirect('dividenda_list')
    else:
        form = DividendaForm(instance=dividenda)
    return render(request, 'akcie/dividenda_form.html', {'form': form})

@login_required
@admin_required
def dividenda_delete(request, pk):
    dividenda = get_object_or_404(Dividenda, pk=pk)
    if request.method == 'POST':
        dividenda.delete()
        return redirect('dividenda_list')
    return render(request, 'akcie/dividenda_confirm_delete.html', {'dividenda': dividenda})

def import_akcie_csv(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Soubor musí být ve formátu CSV.')
            return render(request, 'akcie/import_form.html')

        decoded_file = csv_file.read().decode('utf-8').splitlines()
        reader = csv.reader(decoded_file)
        next(reader)  # Přeskočit hlavičku
        for row in reader:
            # Očekáváme: nazev, pocet_ks, cena_za_kus, hodnota, nakup, zisk_ztráta, ticker, mena
            nazev = row[0]
            pocet_ks = int(row[1])
            cena_za_kus = float(row[2])
            hodnota = float(row[3])
            nakup = float(row[4])
            zisk_ztráta = float(row[5])
            ticker = row[6] if len(row) > 6 else None
            mena = row[7] if len(row) > 7 else 'CZK'
            # Přepočet na CZK pokud není CZK
            cena_za_kus_czk = convert_to_czk(cena_za_kus, mena)
            hodnota_czk = convert_to_czk(hodnota, mena)
            nakup_czk = convert_to_czk(nakup, mena)
            zisk_ztráta_czk = convert_to_czk(zisk_ztráta, mena)
            Akcie.objects.create(
                nazev=nazev,
                pocet_ks=pocet_ks,
                cena_za_kus=cena_za_kus_czk,
                hodnota=hodnota_czk,
                nakup=nakup_czk,
                zisk_ztráta=zisk_ztráta_czk,
                ticker=ticker,
                mena=mena
            )
        messages.success(request, 'Data byla úspěšně importována.')
        return render(request, 'akcie/import_form.html')
    return render(request, 'akcie/import_form.html')

def import_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Očekáváme: nazev, pocet_ks, cena_za_kus, hodnota, nakup, zisk_ztráta, dividenda, ticker, mena
            nazev = row[0]
            pocet_ks = int(row[1])
            cena_za_kus = float(row[2])
            hodnota = float(row[3])
            nakup = float(row[4])
            zisk_ztráta = float(row[5])
            dividenda = float(row[6])
            ticker = row[7] if len(row) > 7 and row[7] else None
            mena = row[8] if len(row) > 8 and row[8] else 'CZK'
            cena_za_kus_czk = convert_to_czk(cena_za_kus, mena)
            hodnota_czk = convert_to_czk(hodnota, mena)
            nakup_czk = convert_to_czk(nakup, mena)
            zisk_ztráta_czk = convert_to_czk(zisk_ztráta, mena)
            dividenda_czk = convert_to_czk(dividenda, mena)
            Akcie.objects.create(
                nazev=nazev,
                pocet_ks=pocet_ks,
                cena_za_kus=cena_za_kus_czk,
                hodnota=hodnota_czk,
                nakup=nakup_czk,
                zisk_ztráta=zisk_ztráta_czk,
                dividenda=dividenda_czk,
                ticker=ticker,
                mena=mena
            )
        log_aktivita("Import dat z Excelu", request.user.username if request.user.is_authenticated else "Anonymní")
        return redirect('akcie_list')
    return render(request, 'akcie/import_form.html')

def export_akcie_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="akcie.csv"'

    writer = csv.writer(response)
    writer.writerow(['Název', 'Počet kusů', 'Cena za kus', 'Hodnota', 'Nákup', 'Zisk/Ztráta'])

    akcie = Akcie.objects.all()
    for a in akcie:
        writer.writerow([
            a.nazev,
            a.pocet_ks,
            f"{a.cena_za_kus:,.2f} Kč",
            f"{a.hodnota:,.2f} Kč",
            f"{a.nakup:,.2f} Kč",
            f"{a.zisk_ztráta:,.2f} Kč"
        ])

    # Luxusní podpis a vodoznak
    writer.writerow([])
    writer.writerow(["--- Luxusní export pro Finanční Poradce Premium ---"])
    return response

@login_required
def export_akcie_json(request):
    """
    Exportuje data o akciích do JSON formátu.
    """
    akcie = Akcie.objects.all().values()
    return JsonResponse(list(akcie), safe=False)

def export_transakce_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="transakce.csv"'

    writer = csv.writer(response)
    writer.writerow(['Akcie', 'Datum', 'Typ', 'Množství', 'Cena'])

    transakce = Transakce.objects.all()
    for t in transakce:
        writer.writerow([t.akcie.nazev, t.datum, t.typ, t.mnozstvi, t.cena])

    return response

def export_dividendy_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="dividendy.csv"'

    writer = csv.writer(response)
    writer.writerow(['Akcie', 'Datum', 'Částka'])

    dividendy = Dividenda.objects.all()
    for d in dividendy:
        writer.writerow([d.akcie.nazev, d.datum, d.castka])

    return response

@login_required
def export_all_data_zip(request):
    """
    Exportuje všechna data (akcie, transakce, dividendy) do jednoho ZIP souboru.
    """
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w') as zip_file:
        # Export akcií
        akcie_data = Akcie.objects.all().values()
        akcie_csv = 'akcie.csv'
        zip_file.writestr(akcie_csv, '\n'.join([','.join(map(str, row.values())) for row in akcie_data]))

        # Export transakcí
        transakce_data = Transakce.objects.all().values()
        transakce_csv = 'transakce.csv'
        zip_file.writestr(transakce_csv, '\n'.join([','.join(map(str, row.values())) for row in transakce_data]))

        # Export dividend
        dividendy_data = Dividenda.objects.all().values()
        dividendy_csv = 'dividendy.csv'
        zip_file.writestr(dividendy_csv, '\n'.join([','.join(map(str, row.values())) for row in dividendy_data]))

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="vsechna_data.zip"'
    return response

def export_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="akcie_export.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Akcie"

    # Přidání záhlaví
    headers = ["Název", "Počet kusů", "Cena za kus", "Hodnota", "Nákup", "Zisk/Ztráta"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="007bff")
    ws.sheet_properties.tabColor = "007bff"

    # Přidání dat
    for akcie in Akcie.objects.all():
        ws.append([
            akcie.nazev,
            akcie.pocet_ks,
            float(akcie.cena_za_kus),
            float(akcie.hodnota),
            float(akcie.nakup),
            float(akcie.zisk_ztráta)
        ])

    # Luxusní branding
    ws.append([])
    ws.append(["--- Luxusní export pro Finanční Poradce Premium ---"])

    log_aktivita("Export dat do Excelu", request.user.username if request.user.is_authenticated else "Anonymní")
    wb.save(response)
    return response

def generate_akcie_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="akcie_report.pdf"'

    p = canvas.Canvas(response)
    add_luxury_branding(p)
    p.setFont("Helvetica", 12)
    p.drawString(100, 800, "Report o akciích")
    y = 780

    akcie = Akcie.objects.all()
    for a in akcie:
        p.drawString(100, y, f"Název: {a.nazev}, Počet kusů: {a.pocet_ks}, Cena za kus: {a.cena_za_kus:,.2f} Kč, Hodnota: {a.hodnota:,.2f} Kč, Zisk/Ztráta: {a.zisk_ztráta:,.2f} Kč, Dividenda: {a.dividenda:,.2f} Kč")
        y -= 20
        if y < 50:
            p.showPage()
            add_luxury_branding(p)
            p.setFont("Helvetica", 12)
            y = 800
    p.showPage()
    p.save()
    return response

def generate_transakce_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transakce_report.pdf"'

    p = canvas.Canvas(response)
    p.setFont("Helvetica", 12)

    p.drawString(100, 800, "Report o transakcích")
    y = 780

    transakce = Transakce.objects.all()
    for t in transakce:
        p.drawString(100, y, f"Akcie: {t.akcie.nazev}, Typ: {t.typ}, Množství: {t.mnozstvi}, Cena: {t.cena}")
        y -= 20

    p.showPage()
    p.save()
    return response

def generate_dividenda_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="dividenda_report.pdf"'

    p = canvas.Canvas(response)
    p.setFont("Helvetica", 12)

    p.drawString(100, 800, "Report o dividendách")
    y = 780

    dividendy = Dividenda.objects.all()
    for d in dividendy:
        p.drawString(100, y, f"Akcie: {d.akcie.nazev}, Částka: {d.castka}, Datum: {d.datum}")
        y -= 20

    p.showPage()
    p.save()
    return response

def export_dashboard_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="dashboard_report.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    add_luxury_branding(p)
    p.setFont("Helvetica", 12)

    p.drawString(100, 750, "Dashboard Report")

    total_akcie = Akcie.objects.count()
    total_transakce = Transakce.objects.count()
    total_dividendy = Dividenda.objects.aggregate(Sum('castka'))['castka__sum'] or 0

    p.drawString(100, 700, f"Počet akcií: {total_akcie}")
    p.drawString(100, 680, f"Počet transakcí: {total_transakce}")
    p.drawString(100, 660, f"Celková částka dividend: {total_dividendy} Kč")

    p.showPage()
    p.save()
    return response

def export_dashboard_graphs_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="dashboard_graphs.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    p.setFont("Helvetica", 12)

    p.drawString(100, 750, "Dashboard Grafy")

    # Screenshot grafů pomocí Selenium
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        driver.get(request.build_absolute_uri(reverse('dashboard')))
        driver.set_window_size(1200, 800)
        screenshot = driver.get_screenshot_as_png()
        image = ImageReader(BytesIO(screenshot))
        p.drawImage(image, 50, 400, width=500, height=300)
    finally:
        driver.quit()

    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='dashboard_graphs.pdf')

def dashboard(request):
    query = request.GET.get('q')

    if query:
        akcie = Akcie.objects.filter(Q(nazev__icontains=query))
    else:
        akcie = Akcie.objects.all()

    # Přepočet všech akcií na CZK (pro jistotu i při starších datech/importech)
    total_akcie = akcie.count()
    total_hodnota = 0
    total_zisk_ztrata = 0
    akcie_data = []
    for a in akcie:
        # Pokud by model měl pole 'mena', použij ho, jinak předpokládej CZK
        mena = getattr(a, 'mena', 'CZK')
        hodnota_czk = convert_to_czk(a.hodnota, mena)
        zisk_ztrata_czk = convert_to_czk(a.zisk_ztráta, mena)
        total_hodnota += hodnota_czk
        akcie_data.append({'nazev': a.nazev, 'hodnota': hodnota_czk})

    transakce_monthly = (
        Transakce.objects.annotate(month=TruncMonth('datum'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    dividendy_data = (
        Dividenda.objects.values('akcie__nazev')
        .annotate(total_castka=Sum('castka'))
        .order_by('akcie__nazev')
    )

    current_year = now().year
    months = ["Leden", "Únor", "Březen", "Duben", "Květen", "Červen", "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]
    investment_data = []

    for month in range(1, 13):
        total_investment = Transakce.objects.filter(datum__year=current_year, datum__month=month).aggregate(Sum('cena'))['cena__sum'] or 0
        investment_data.append(total_investment)

    transakce_typy = (
        Transakce.objects.values('typ')
        .annotate(count=Count('id'))
        .order_by('typ')
    )
    context = {
        'total_akcie': total_akcie,
        'total_hodnota': total_hodnota,
        'total_zisk_ztrata': total_zisk_ztrata,
        'akcie_data': list(akcie_data),
        'transakce_monthly': list(transakce_monthly),
        'dividendy_data': list(dividendy_data),
        'months': months,
        'investment_data': investment_data,
        'query': query,
        'transakce_typy': list(transakce_typy),
    }
    return render(request, 'akcie/dashboard.html', context)

def aktivity_list(request):
    query_user = request.GET.get('user')
    query_action = request.GET.get('action')

    aktivity = Aktivita.objects.all()

    if query_user:
        aktivity = aktivity.filter(uzivatel__icontains=query_user)
    if query_action:
        aktivity = aktivity.filter(akce__icontains=query_action)

    aktivity = aktivity.order_by('-datum_cas')

    return render(request, 'akcie/aktivity_list.html', {
        'aktivity': aktivity,
        'query_user': query_user,
        'query_action': query_action
    })

def export_aktivity_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="aktivity.csv"'

    writer = csv.writer(response)
    writer.writerow(['Akce', 'Uživatel', 'Datum a čas'])

    for aktivita in Aktivita.objects.all():
        writer.writerow([aktivita.akce, aktivita.uzivatel, aktivita.datum_cas])

    return response

def export_aktivity_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="aktivity_report.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica", 12)

    p.drawString(100, 750, "Seznam aktivit")

    y = 720
    for aktivita in Aktivita.objects.all():
        p.drawString(100, y, f"Akce: {aktivita.akce}, Uživatel: {aktivita.uzivatel}, Datum a čas: {aktivita.datum_cas}")
        y -= 20
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 12)
            y = 750

    p.save()
    return response

@login_required
@admin_required
def auditlog_list(request):
    logs = AuditLog.objects.order_by('-timestamp')[:200]
    return render(request, 'akcie/auditlog_list.html', {'logs': logs})

def send_monthly_report():
    users = CustomUser.objects.filter(receive_monthly_reports=True)
    pdf_path = generate_pdf_report()  # Generování PDF reportu
    for user in users:
        subject = "Měsíční report investic"
        body = f"Dobrý den, {user.username},\n\nPřikládáme měsíční report vašich investic."
        email = EmailMessage(
            subject,
            body,
            'your_email@gmail.com',  # Odesílatel
            [user.email_for_reports]  # Příjemce
        )

        email.attach_file(pdf_path)
        email.send()

def generate_pdf_report():
    pdf_path = os.path.join('staticfiles', 'reports', 'monthly_report.pdf')
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    c = canvas.Canvas(pdf_path)
    c.drawString(100, 750, "Měsíční report investic")
    c.drawString(100, 730, "Tento report obsahuje souhrn vašich investic za poslední měsíc.")
    c.save()
    return pdf_path

@login_required
def user_preferences(request):
    if request.method == 'POST':
        form = CustomUserForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = CustomUserForm(instance=request.user)
    return render(request, 'akcie/user_preferences.html', {'form': form})

def export_hot_investments_csv(request):
    """Exportuje seznam hot investic do CSV."""
    hot_investments = fetch_hot_investments()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="hot_investments.csv"'

    writer = csv.writer(response)
    writer.writerow(['Název', 'Cena za kus', 'Hodnota'])
    for investment in hot_investments:
        writer.writerow([investment['nazev'], investment['cena_za_ks'], investment['hodnota']])

    return response

def search_stocks(request):
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse([], safe=False)
    try:
        stock = yf.Ticker(query)
        info = stock.info
        result = [{
            'nazev': info.get('shortName', query),
            'ticker': query,
            'cena': info.get('regularMarketPrice', 'N/A'),
            'mena': info.get('currency', 'CZK')
        }]
        return JsonResponse(result, safe=False)
    except Exception as e:
        print(f"Chyba při hledání akcie podle tickeru: {e}")

    try:
        # Search by stock name (example: broader search logic)
        tickers = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA']  # Example list of tickers
        results = []
        for ticker in tickers:
            stock = yf.Ticker(ticker)
            info = stock.info
            if query.lower() in info.get('shortName', '').lower():
                results.append({
                    'nazev': info.get('shortName', ticker),
                    'ticker': ticker,
                    'cena': info.get('regularMarketPrice', 'N/A')
                })
        return JsonResponse(results, safe=False)
    except Exception as e:
        print(f"Chyba při hledání akcie podle názvu: {e}")
        return JsonResponse([], safe=False)

def add_stock(request):
    ticker = request.GET.get('ticker', '')
    if not ticker:
        return JsonResponse({'error': 'Ticker is required'}, status=400)

    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        Akcie.objects.create(
            nazev=info.get('shortName', ticker),
            pocet_ks=0,  # Default value, can be updated later
            cena_za_kus=info.get('regularMarketPrice', 0),
            hodnota=0,  # Default value
            nakup=0,  # Default value
            zisk_ztráta=0,  # Default value
            dividenda=0  # Default value
        )
        return JsonResponse({'message': 'Akcie byla úspěšně přidána!'}, status=201)
    except Exception as e:
        print(f"Chyba při přidávání akcie: {e}")
        return JsonResponse({'error': 'Došlo k chybě při přidávání akcie.'}, status=500)

def history_dates(request):
    ticker = request.GET.get('ticker')
    if not ticker:
        return JsonResponse([], safe=False)
    try:
        stock = yf.Ticker(ticker)
        hist = stock.history(period='max')
        dates = [str(d.date()) for d in hist.index]
        # Vrátíme pouze unikátní datumy (např. 1x za den)
        return JsonResponse(sorted(list(set(dates))), safe=False)
    except Exception as e:
        print(f"Chyba při získávání historických dat: {e}")
        return JsonResponse([], safe=False)

@user_passes_test(is_admin)
def admin_only_view(request):
    # ... pouze pro adminy ...
    pass

@user_passes_test(is_poradce)
def poradce_only_view(request):
    # ... pouze pro poradce ...
    pass

@user_passes_test(is_klient)
def klient_only_view(request):
    # ... pouze pro klienty ...
    pass

@csrf_exempt
def api_akcie_list(request):
    if request.method == 'GET':
        data = list(Akcie.objects.values())
        return JsonResponse(data, safe=False)

@csrf_exempt
def api_transakce_list(request):
    if request.method == 'GET':
        data = list(Transakce.objects.values())
        return JsonResponse(data, safe=False)

@csrf_exempt
def api_dividenda_list(request):
    if request.method == 'GET':
        data = list(Dividenda.objects.values())
        return JsonResponse(data, safe=False)

def klienti(request):
    return render(request, 'akcie/klienti.html')

@login_required
@poradce_required
def analyzy(request):
    # Žebříček portfolií podle hodnoty (součet hodnoty všech akcií v portfoliu)
    zebricek = []
    for klient in Klient.objects.all():
        for portfolio in klient.portfolia.all():
            # Zde lze napojit na model Akcie, pokud bude napojení na portfolio
            hodnota = 0  # Zatím placeholder, v budoucnu součet hodnoty akcií v portfoliu
            zebricek.append({
                'klient': f"{klient.jmeno} {klient.prijmeni}",
                'portfolio': portfolio.nazev,
                'hodnota': hodnota
            })
    zebricek = sorted(zebricek, key=lambda x: x['hodnota'], reverse=True)[:10]

    # AI predikce (velmi jednoduchá, lineární trend na základě hodnoty portfolia)
    predikce = []
    for item in zebricek:
        # Odhad růstu o 2 % za měsíc (placeholder)
        predikce.append({
            'portfolio': item['portfolio'],
            'predikce': item['hodnota'] * 1.02
        })
    return render(request, 'akcie/analyzy.html', {'zebricek': zebricek, 'predikce': predikce})

def reporty(request):
    return render(request, 'akcie/reporty.html')

def analyzy(request):
    return render(request, 'akcie/analyzy.html')

class Notifikace:
    def __init__(self, typ, zprava, datum):
        self.typ = typ
        self.zprava = zprava
        self.datum = datum

@login_required
def upozorneni(request):
    # Zatím pouze ukázkové notifikace, v budoucnu napojení na model a alerty
    notifikace = [
        Notifikace('Systém', 'Byl vygenerován nový měsíční report.', timezone.now()),
        Notifikace('Portfolio', 'Vaše portfolio překročilo hodnotu 1 000 000 Kč.', timezone.now()),
        Notifikace('Upozornění', 'Blíží se výplata dividendy.', timezone.now()),
    ]
    return render(request, 'akcie/upozorneni.html', {'notifikace': notifikace})

def nastaveni(request):
    return render(request, 'akcie/nastaveni.html')

def vip(request):
    return render(request, 'akcie/vip.html')

def chat(request):
    return render(request, 'akcie/chat.html')

@login_required
def integrace(request):
    messages = []
    if request.method == 'POST' and request.FILES.get('import_file'):
        import_file = request.FILES['import_file']
        try:
            if import_file.name.endswith('.csv'):
                df = pd.read_csv(import_file)
            elif import_file.name.endswith('.xlsx'):
                df = pd.read_excel(import_file)
            else:
                messages.append('Nepodporovaný formát souboru.')
                return render(request, 'akcie/integrace.html', {'messages': messages})
            # Ukázka: výpis prvních 3 řádků
            messages.append(f'Načteno {len(df)} záznamů. Náhled:')
            for i, row in df.head(3).iterrows():
                messages.append(str(row.to_dict()))
            # Zde lze napojit na import transakcí/portfolia
        except Exception as e:
            messages.append(f'Chyba při importu: {e}')
    return render(request, 'akcie/integrace.html', {'messages': messages})

@login_required
@poradce_required
def poradce_dashboard(request):
    # Zobrazí přehled klientů a jejich portfolií pro přihlášeného poradce
    klienti = Klient.objects.filter(poradce=request.user)
    klient_portfolia = []
    for klient in klienti:
        portfolia = klient.portfolia.all()
        klient_portfolia.append({
            'klient': klient,
            'portfolia': portfolia,
        })
    return render(request, 'akcie/poradce_dashboard.html', {'klient_portfolia': klient_portfolia})
